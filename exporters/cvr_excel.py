"""
Export CVR → Excel — snapshot periode (Value/Cost/Margin/Forecast).
"""
import os
from datetime import datetime


def generate_cvr_excel(project, period) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    from models import CvrCommentary

    wb = Workbook()
    ws = wb.active
    ws.title = f"CVR {period.periode}"

    BLUE = "1F3864"
    money_fmt = "#,##0"
    thin = lambda color="CCCCCC": Border(
        left=Side(style="thin", color=color), right=Side(style="thin", color=color),
        top=Side(style="thin", color=color), bottom=Side(style="thin", color=color))

    r = 1
    ws.merge_cells(f"A{r}:K{r}")
    ws.cell(r, 1, f"CVR {project.nama} — {period.periode} ({period.status})").font = Font(bold=True, size=14, color=BLUE)
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    ws.cell(r, 1, f"Update: {datetime.now().strftime('%Y-%m-%d %H:%M')} · disusun {period.disusun_oleh}").font = Font(italic=True, size=10, color="888888")
    r += 2

    headers = ["Kode", "Uraian", "Value Cert.", "Value Int.", "Cost Actual",
               "Accrual", "Committed Out.", "CTC", "Metode", "FC Cost", "FC Value"]
    widths = [10, 40, 14, 14, 14, 12, 15, 13, 12, 14, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(r, i, h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        cell.border = thin("FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    r += 1

    totals = {k: 0.0 for k in ["value_certified", "value_internal", "cost_actual",
                               "cost_accrual", "cost_committed_outstanding",
                               "forecast_cost_to_complete", "forecast_final_cost",
                               "forecast_final_value"]}
    for line in period.lines:
        it = line.rap_item
        ws.cell(r, 1, it.kode_rap if it else "")
        ws.cell(r, 2, it.uraian_baku if it else "—")
        ws.cell(r, 3, line.value_certified).number_format = money_fmt
        ws.cell(r, 4, line.value_internal).number_format = money_fmt
        ws.cell(r, 5, line.cost_actual).number_format = money_fmt
        ws.cell(r, 6, line.cost_accrual).number_format = money_fmt
        ws.cell(r, 7, line.cost_committed_outstanding).number_format = money_fmt
        ws.cell(r, 8, line.forecast_cost_to_complete).number_format = money_fmt
        ws.cell(r, 9, line.metode_ctc)
        ws.cell(r, 10, line.forecast_final_cost).number_format = money_fmt
        ws.cell(r, 11, line.forecast_final_value).number_format = money_fmt
        for c in range(1, 12):
            ws.cell(r, c).border = thin()
            ws.cell(r, c).font = Font(size=9)
        for k in totals:
            totals[k] += getattr(line, k)
        r += 1

    # Total
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1, "TOTAL").font = Font(bold=True, size=10)
    for i, k in enumerate(["value_certified", "value_internal", "cost_actual",
                           "cost_accrual", "cost_committed_outstanding",
                           "forecast_cost_to_complete", "forecast_final_cost",
                           "forecast_final_value"]):
        ws.cell(r, i + 3, totals[k]).number_format = money_fmt
    for c in range(1, 12):
        ws.cell(r, c).font = Font(bold=True, size=10)
        ws.cell(r, c).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        ws.cell(r, c).border = Border(top=Side(style="double"), bottom=Side(style="double"))
    r += 2

    # Commentary
    ws.cell(r, 1, "Catatan CVR").font = Font(bold=True, size=11)
    r += 1
    for c in CvrCommentary.query.filter_by(cvr_period_id=period.id).all():
        ws.cell(r, 1, f"{c.penyusun}: {c.teks}").font = Font(size=9, italic=True)
        r += 1

    out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                            "data", f"cvr_export_{period.id}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
