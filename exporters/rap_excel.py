"""
Export RAP → Excel — format mudah dibaca manual.
Wajib ada: kalau sistem berhenti dipakai, data tetap bisa dibaca & dilanjutkan.
"""
import os
from datetime import datetime


def generate_rap_excel(project, version) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    from models import RapItem, PrelimItem, RiskAllowance

    wb = Workbook()

    # ── Sheet 1: RAP items ──
    ws = wb.active
    ws.title = "RAP"

    BLUE = "1F3864"
    money_fmt = "#,##0"
    thin = lambda color="CCCCCC": Border(
        left=Side(style="thin", color=color), right=Side(style="thin", color=color),
        top=Side(style="thin", color=color), bottom=Side(style="thin", color=color))

    r = 1
    ws.merge_cells(f"A{r}:L{r}")
    ws.cell(r, 1, f"RAP {project.nama} — {version.versi} ({version.status})").font = Font(bold=True, size=14, color=BLUE)
    r += 1
    ws.merge_cells(f"A{r}:L{r}")
    ws.cell(r, 1, f"Update: {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
                  f"{version.catatan_revisi or version.disusun_oleh or ''}").font = Font(italic=True, size=10, color="888888")
    r += 2

    headers = ["Kode", "Uraian", "Jenis Biaya", "Satuan", "Vol BOQ", "Faktor",
               "Vol RAP", "Harga Satuan", "Total RAP", "Terikat", "Sisa", "Sumber Harga"]
    widths = [10, 45, 12, 8, 12, 8, 12, 15, 15, 15, 15, 13]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(r, i, h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        cell.border = thin("FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    r += 1

    items = RapItem.query.filter_by(project_id=project.id, rap_version_id=version.id) \
                         .order_by(RapItem.kode_rap).all()
    total_rap = 0.0
    total_terikat = 0.0
    for it in items:
        ws.cell(r, 1, it.kode_rap)
        ws.cell(r, 2, it.uraian_baku)
        ws.cell(r, 3, it.jenis_biaya)
        ws.cell(r, 4, it.satuan)
        ws.cell(r, 5, it.vol_boq).number_format = money_fmt
        ws.cell(r, 6, it.faktor)
        ws.cell(r, 7, it.vol_rap).number_format = money_fmt
        ws.cell(r, 8, it.hsat_rap).number_format = money_fmt
        ws.cell(r, 9, it.total_rap).number_format = money_fmt
        ws.cell(r, 10, it.terikat).number_format = money_fmt
        ws.cell(r, 11, it.sisa_budget).number_format = money_fmt
        ws.cell(r, 12, it.sumber_harga)
        if it.is_consumable:
            ws.cell(r, 2).value += "  [HABIS PAKAI]"
            for c in range(1, 13):
                ws.cell(r, c).font = Font(size=9, italic=True, color="888888")
        for c in range(1, 13):
            ws.cell(r, c).border = thin()
            ws.cell(r, c).font = Font(size=9)
        total_rap += it.total_rap
        total_terikat += it.terikat
        r += 1

    # Grand total
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1, "TOTAL").font = Font(bold=True, size=10)
    ws.cell(r, 9, total_rap).number_format = money_fmt
    ws.cell(r, 10, total_terikat).number_format = money_fmt
    ws.cell(r, 11, total_rap - total_terikat).number_format = money_fmt
    for c in range(1, 13):
        ws.cell(r, c).font = Font(bold=True, size=10)
        ws.cell(r, c).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        ws.cell(r, c).border = Border(top=Side(style="double"), bottom=Side(style="double"))

    # ── Sheet 2: Prelim ──
    ws2 = wb.create_sheet("Prelim")
    ws2.cell(1, 1, "Uraian").font = Font(bold=True)
    ws2.cell(1, 2, "Biaya/Bulan").font = Font(bold=True)
    ws2.cell(1, 3, "Durasi (bln)").font = Font(bold=True)
    ws2.cell(1, 4, "Total").font = Font(bold=True)
    pre = PrelimItem.query.filter_by(project_id=project.id, rap_version_id=version.id).all()
    rr = 2
    for p in pre:
        ws2.cell(rr, 1, p.uraian)
        ws2.cell(rr, 2, p.biaya_per_bulan).number_format = money_fmt
        ws2.cell(rr, 3, p.durasi_rencana_bulan)
        ws2.cell(rr, 4, p.total).number_format = money_fmt
        rr += 1

    # ── Sheet 3: Risk Allowance ──
    ws3 = wb.create_sheet("Risk Allowance")
    ws3.cell(1, 1, "Nama").font = Font(bold=True)
    ws3.cell(1, 2, "Nilai").font = Font(bold=True)
    ws3.cell(1, 3, "Pemicu").font = Font(bold=True)
    ws3.cell(1, 4, "Status").font = Font(bold=True)
    risks = RiskAllowance.query.filter_by(project_id=project.id, rap_version_id=version.id).all()
    rr = 2
    for rk in risks:
        ws3.cell(rr, 1, rk.nama)
        ws3.cell(rr, 2, rk.nilai).number_format = money_fmt
        ws3.cell(rr, 3, rk.pemicu)
        ws3.cell(rr, 4, rk.status)
        rr += 1

    out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                            "data", f"rap_export_{project.id}_{version.id}.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
