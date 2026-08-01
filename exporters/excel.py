"""
Export Excel — generate rekapitulasi pembayaran subkon.
"""

import os
from datetime import datetime


def generate_excel(subcons: list) -> str:
    """Generate rekapitulasi pembayaran. `subcons` = list objek Vendor."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring"

    BLUE = "1F3864"
    LIGHT_BLUE = "D6E4F0"
    YELLOW = "FFF2CC"
    GREEN = "E2EFDA"
    money_fmt = '#,##0'
    pct_fmt = '0.00%'

    thin = lambda color="CCCCCC": Border(
        left=Side(style='thin', color=color),
        right=Side(style='thin', color=color),
        top=Side(style='thin', color=color),
        bottom=Side(style='thin', color=color)
    )

    r = 1
    # Title
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1, "REKAPITULASI PEMBAYARAN SUB KONTRAKTOR").font = Font(bold=True, size=14, color=BLUE)
    r += 1
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1, f"Update: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10, color="888888")
    r += 2

    # Headers
    headers = ["No", "Sub Kontraktor / Pekerjaan", "Nilai Kontrak",
               "Total Dibayar", "Retensi", "Sisa Bayar", "Progress %", "Status"]
    widths = [5, 55, 18, 18, 15, 18, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(r, i, h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        cell.border = thin("FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 30
    r += 1

    grand = {"contract": 0, "paid": 0, "retention": 0}
    idx = 0

    for subcon in sorted(subcons, key=lambda s: s.name.lower()):
        if subcon.total_contract <= 0:
            continue
        idx += 1
        grand["contract"] += subcon.total_contract
        grand["paid"] += subcon.total_paid
        grand["retention"] += subcon.total_retention

        # Subcon row
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(r, 1, f"{idx}. {subcon.name}").font = Font(bold=True, size=10, color=BLUE)
        ws.cell(r, 3, subcon.total_contract).number_format = money_fmt
        ws.cell(r, 4, subcon.total_paid).number_format = money_fmt
        ws.cell(r, 5, subcon.total_retention).number_format = money_fmt
        ws.cell(r, 6, max(subcon.total_contract - subcon.total_paid - subcon.total_retention, 0)).number_format = money_fmt
        ws.cell(r, 7, subcon.pct_vs_contract / 100).number_format = pct_fmt
        ws.cell(r, 8, "LUNAS" if subcon.is_lunas else "PROSES")
        for c in range(1, 9):
            ws.cell(r, c).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
            ws.cell(r, c).border = thin()
        r += 1

    # Grand total
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1, "TOTAL KESELURUHAN").font = Font(bold=True, size=11)
    ws.cell(r, 3, grand["contract"]).number_format = money_fmt
    ws.cell(r, 4, grand["paid"]).number_format = money_fmt
    ws.cell(r, 5, grand["retention"]).number_format = money_fmt
    ws.cell(r, 6, max(grand["contract"] - grand["paid"] - grand["retention"], 0)).number_format = money_fmt
    pct = grand["paid"] / grand["contract"] if grand["contract"] > 0 else 0
    ws.cell(r, 7, pct).number_format = pct_fmt
    for c in range(1, 9):
        ws.cell(r, c).font = Font(bold=True, size=11)
        ws.cell(r, c).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        ws.cell(r, c).border = Border(
            top=Side(style="double"), bottom=Side(style="double")
        )

    out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                            "data", "monitoring_export.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def _status(pct):
    if pct >= 100: return "LUNAS"
    if pct >= 80:  return "On Track"
    if pct >= 50:  return "Progress"
    return "Awal"
