"""
Sertifikat Parser — baca file .xlsx Sertifikat Pembayaran.
Cell coordinates dikonfigurasi via config.py, bukan hardcoded.
"""

import os
import re
from datetime import datetime


def parse_sertifikat(filepath: str) -> dict:
    """
    Parse satu file sertifikat pembayaran.
    Returns dict dengan semua data yang berhasil diekstrak,
    atau raises ValueError dengan pesan yang jelas jika file tidak valid.
    """
    import openpyxl
    from config import SERTIFIKAT_CELLS, SP_SHEET_PREFIX

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ValueError(f"File tidak bisa dibuka: {e}")

    # Cari SP sheets
    sp_sheets = [s for s in wb.sheetnames if s.strip().upper().startswith(SP_SHEET_PREFIX)]
    if not sp_sheets:
        raise ValueError(f"Tidak ada sheet '{SP_SHEET_PREFIX}*' di file ini. "
                         f"Sheet yang ada: {', '.join(wb.sheetnames)}")

    results = []
    parse_errors = []

    for sheet_name in sp_sheets:
        ws = wb[sheet_name]
        try:
            data = _parse_single_sheet(ws, sheet_name, SERTIFIKAT_CELLS)
            results.append(data)
        except Exception as e:
            parse_errors.append({"sheet": sheet_name, "error": str(e)})

    if not results:
        raise ValueError(f"Semua sheet gagal di-parse. Errors: {parse_errors}")

    # Sort by payment number
    results.sort(key=lambda x: x.get("payment_num", 0))

    return {
        "filename": os.path.basename(filepath),
        "payments": results,
        "parse_errors": parse_errors,   # sheet yang gagal tapi tidak blocking
    }


def _parse_single_sheet(ws, sheet_name: str, cells: dict) -> dict:
    """Parse satu SP sheet."""

    # Validasi: harus ada kata "SERTIFIKAT" di header cell
    r, c = cells["header_check"]
    header_val = str(ws.cell(r, c).value or "")
    if "SERTIFIKAT" not in header_val.upper():
        raise ValueError(
            f"Cell ({r},{c}) tidak mengandung 'SERTIFIKAT'. "
            f"Nilai: '{header_val}'. Cek apakah ini benar-benar Sertifikat Pembayaran."
        )

    data = {"_sheet": sheet_name}

    # Subcon name
    r, c = cells["subcon_name"]
    data["subcon_name"] = _str(ws.cell(r, c).value)
    if not data["subcon_name"]:
        raise ValueError(f"Nama subkon kosong di cell ({r},{c})")

    # Payment number
    r, c = cells["payment_number"]
    pay_num_raw = _str(ws.cell(r, c).value)
    data["payment_number_raw"] = pay_num_raw
    match = re.search(r'(\d+)', pay_num_raw)
    data["payment_num"] = int(match.group(1)) if match else 0

    # Work description
    r, c = cells["work_desc"]
    data["work_desc"] = _str(ws.cell(r, c).value)

    # Date
    r, c = cells["date"]
    data["date"] = _parse_date(ws.cell(r, c).value)

    # SPK number
    r, c = cells["spk_number"]
    data["spk_number"] = _str(ws.cell(r, c).value)

    # Contract value — selalu pake Nilai SPK Awal (R16C12) yang tetap per SPK
    r, c = cells["contract_initial"]
    data["contract_value"] = _float(ws.cell(r, c).value)

    # Contract variation — scan semua row untuk penambahan/pengurangan SPK/PO
    total_add = 0
    total_reduce = 0
    for vr in range(18, 50):
        label = _str(ws.cell(vr, 2).value).lower()
        if "penambahan spk" in label or "penambahan po" in label:
            v = _float(ws.cell(vr, 14).value) or _float(ws.cell(vr, 9).value)
            total_add += abs(v)
        elif "pengurangan spk" in label or "pengurangan po" in label:
            v = _float(ws.cell(vr, 14).value) or _float(ws.cell(vr, 9).value)
            if v == 0:
                v = _float(ws.cell(vr, 12).value)
            total_reduce += abs(v)
        elif "prestasi pengiriman" in label or "sub total" in label:
            break
    data["total_penambahan"] = total_add
    data["total_pengurangan"] = total_reduce

    # DP / Uang Muka
    r, c = cells["dp_amount"]
    data["dp_amount"] = _float(ws.cell(r, c).value)

    # Retention pct
    r, c = cells["retention_pct"]
    ret_raw = _float(ws.cell(r, c).value)
    # Kalau nilainya < 1, kemungkinan sudah dalam desimal (0.05 = 5%)
    data["retention_pct"] = ret_raw * 100 if ret_raw < 1 else ret_raw
    if data["retention_pct"] <= 0:
        from config import DEFAULT_RETENTION_PCT
        data["retention_pct"] = DEFAULT_RETENTION_PCT

    # Cumulative progress — cari baris "Nilai SPK/PO yang telah dilaksanakan"
    data["cumulative_progress"] = 0.0
    for sr in range(20, 50):
        txt = _str(ws.cell(sr, 2).value)
        low = txt.lower()
        if ("nilai spk" in low or "nilai po" in low) and "dilaksanakan" in low:
            data["cumulative_progress"] = _float(ws.cell(sr, 7).value)
            break

    # Net payment — cari "Sub Total 3" di C8, fallback ke "Total Pembayaran s/d lalu"
    data["net_payment"] = 0.0
    data["total_invoice"] = 0.0
    found_total = False
    # Primary: cari "Sub Total 3" di C8
    for sr in range(30, 65):
        txt = _str(ws.cell(sr, 8).value)
        if "sub total 3" in txt.lower():
            data["net_payment"] = _float(ws.cell(sr, 14).value)
            # Cek PPN di C8 baris berikutnya
            ppn_txt = _str(ws.cell(sr + 1, 8).value)
            if "ppn" in ppn_txt.lower():
                ppn_val = _float(ws.cell(sr + 1, 14).value)
                data["total_invoice"] = data["net_payment"] + ppn_val
            else:
                data["total_invoice"] = data["net_payment"]
            found_total = True
            break
    if not found_total:
        # Fallback 1: cari "Total Pembayaran s/d lalu" di C2
        for sr in range(30, 65):
            txt = _str(ws.cell(sr, 2).value)
            if "total pembayaran s/d lalu" in txt.lower():
                net3 = _float(ws.cell(sr + 2, 14).value)
                ppn_val = _float(ws.cell(sr + 3, 14).value)
                if ppn_val > 0 and abs(ppn_val - net3) > 1:
                    data["total_invoice"] = net3 + ppn_val
                else:
                    data["total_invoice"] = net3
                data["net_payment"] = net3
                found_total = True
                break
    if not found_total:
        # Fallback 2: cari "Pembayaran Uang Muka" → net di baris sebelumnya
        for sr in range(40, 60):
            txt = _str(ws.cell(sr, 2).value)
            if "pembayaran uang muka" in txt.lower():
                data["net_payment"] = _float(ws.cell(sr - 1, 14).value)
                data["total_invoice"] = data["net_payment"]
                break

    # Payment description
    data["payment_desc"] = f"Pembayaran SP-{data['payment_num']}"
    data["is_dp"] = False

    return data


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace("Rp.", "").replace("Rp", "").strip()
        s = s.replace(",", "").replace(" ", "")
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _parse_date(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Coba parse berbagai format umum
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Kembalikan raw string (max 10 char) sebagai fallback
    return s[:10]
